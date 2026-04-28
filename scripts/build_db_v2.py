"""
v2: Extends build_db.py with ~22 new restaurants, quality scoring, tiers, Top Picks tab.
"""

import re
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from build_db import CAT_A as CAT_A_BASE, CAT_B as CAT_B_BASE, CAT_C as CAT_C_BASE, CAT_D as CAT_D_BASE

OUT = "/sessions/keen-focused-ritchie/mnt/outputs/HK_Vegetarian_Restaurants_Database.xlsx"

# ============================================================
# NEW ADDITIONS (~22 restaurants — duplicates already filtered)
# ============================================================

# --- A. Chinese / Buddhist additions ---
A_NEW = [
    {
        "Name (EN)": "Gaia Veggie Shop",
        "Name (中文)": "—",
        "Sub-type": "Cantonese mock-meat",
        "District": "Tsim Sha Tsui (also CWB)",
        "Address": "Shop 212, 2/F Mira Place 1, 132 Nathan Rd, TST",
        "Phone": "+852 2376 1186",
        "Website / IG": "gaiaveggie.com",
        "Booking": "Phone, OpenRice",
        "Hours": "11:00–22:00",
        "Price (HKD/pp)": "100–180",
        "Veg type": "Vegetarian",
        "Jain": "On request",
        "Signature / sample dishes (with prices where known)": "Veg fried rice; mushroom 'lamb' skewers; mock-meat curry; sweet & sour 'pork'",
        "Google rating": "4.1 (200+)",
        "OpenRice rating": "4.0",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Long-running mock-meat-focused chain, multiple branches.",
        "Sources (URLs)": "OpenRice; Time Out HK",
    },
    {
        "Name (EN)": "Amitabha Buddha Vegetarian",
        "Name (中文)": "—",
        "Sub-type": "Buddhist 齋",
        "District": "Hung Hom",
        "Address": "G/F, 41-43 Station Lane, Hung Hom",
        "Phone": "+852 2215 4882",
        "Website / IG": "—",
        "Booking": "Phone, walk-in",
        "Hours": "Verify (lunch + dinner)",
        "Price (HKD/pp)": "60–120",
        "Veg type": "Vegetarian (no allium)",
        "Jain": "Yes (de facto)",
        "Signature / sample dishes (with prices where known)": "Set lunch; mock-meat curry; veg dim sum",
        "Google rating": "3.8 (40+)",
        "OpenRice rating": "3.9",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "Trusted",
        "Operational status": "PERMANENTLY CLOSED",
        "Notes": "Buddhist temple-affiliated. Phone before going — limited public info on current hours.",
        "Sources (URLs)": "OpenRice; HappyCow",
    },
    {
        "Name (EN)": "Light Vegetarian",
        "Name (中文)": "光線素食",
        "Sub-type": "Cantonese veg dim sum",
        "District": "Jordan",
        "Address": "Near Jordan MTR (verify exact street address)",
        "Phone": "—",
        "Website / IG": "—",
        "Booking": "Walk-in",
        "Hours": "Lunch yum cha; dinner",
        "Price (HKD/pp)": "60–150",
        "Veg type": "Vegetarian",
        "Jain": "No",
        "Signature / sample dishes (with prices where known)": "Lunch yum cha (veg dim sum); set dinner",
        "Google rating": "—",
        "OpenRice rating": "—",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "Verify (low online presence)",
        "Operational status": "Verify",
        "Notes": "Mentioned in HK veg guides but minimal online presence — phone before going.",
        "Sources (URLs)": "Green Queen 2025 mention",
    },
]

# --- B. Modern Plant-Based additions ---
B_NEW = [
    {
        "Name (EN)": "Mayse Artisan Bakery",
        "Name (中文)": "—",
        "Sub-type": "Vegan bakery",
        "District": "Tai Po",
        "Address": "Tai Mei Tuk, Tai Po",
        "Phone": "—",
        "Website / IG": "@mayseartisanbakery",
        "Booking": "Walk-in",
        "Hours": "Verify (weekend-heavy)",
        "Price (HKD/pp)": "60–150",
        "Veg type": "Vegan",
        "Jain": "Partial",
        "Signature / sample dishes (with prices where known)": "Sourdough bread; vegan pastries; sourdough pizza; vegan cakes",
        "Google rating": "4.6 (100+)",
        "OpenRice rating": "—",
        "HappyCow rating": "4.7",
        "Michelin / Tatler / Time Out": "Time Out HK feature",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Latvian family-run; northern European recipes. Quiet location near Plover Cove.",
        "Sources (URLs)": "https://www.timeout.com/hong-kong; Honeycombers",
    },
    {
        "Name (EN)": "Mayse Pizzeria",
        "Name (中文)": "—",
        "Sub-type": "Vegan pizza / sourdough",
        "District": "Jordan",
        "Address": "G/F 5-11 Woosung St, Jordan",
        "Phone": "—",
        "Website / IG": "facebook.com/maysebakery",
        "Booking": "Walk-in",
        "Hours": "Wed-Sun 12:00–21:00 (varies)",
        "Price (HKD/pp)": "80–150",
        "Veg type": "Vegan",
        "Jain": "Partial",
        "Signature / sample dishes (with prices where known)": "Vegan pizzas; panini; sourdough breads",
        "Google rating": "4.3 (100+)",
        "OpenRice rating": "—",
        "HappyCow rating": "4.5",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "Trusted",
        "Operational status": "PERMANENTLY CLOSED (per HappyCow listing)",
        "Notes": "Sister project to Mayse Artisan Bakery. Approx 7 min walk from Jordan MTR.",
        "Sources (URLs)": "HappyCow; Time Out HK",
    },
    {
        "Name (EN)": "VW Vegan Cafe",
        "Name (中文)": "—",
        "Sub-type": "Vegan cafe / desserts",
        "District": "Tai Po",
        "Address": "Shop 227F-06, 2/F Plover Cove Garden Arcade, 3 Plover Cove Rd, Tai Po",
        "Phone": "+852 5466 9854",
        "Website / IG": "facebook.com/v.wvegan",
        "Booking": "Phone",
        "Hours": "Wed-Fri 11:30–18:00; Sat-Sun 11:30–21:00",
        "Price (HKD/pp)": "50–120",
        "Veg type": "Vegan",
        "Jain": "Partial",
        "Signature / sample dishes (with prices where known)": "Vegan lemon donuts; blueberry-cheesecake tarts; mochi brownies; bagels",
        "Google rating": "4.7 (150+)",
        "OpenRice rating": "—",
        "HappyCow rating": "4.7",
        "Michelin / Tatler / Time Out": "Time Out HK + Green Queen",
        "Review-quality flag": "Trusted",
        "Operational status": "PERMANENTLY CLOSED (~Jan 2024 per Facebook page)",
        "Notes": "Newer 2025 vegan dessert cafe. High-rated — fresh signal.",
        "Sources (URLs)": "Time Out HK; Green Queen",
    },
    {
        "Name (EN)": "The Cakery (Lee Garden)",
        "Name (中文)": "—",
        "Sub-type": "Vegan / refined-sugar-free bakery",
        "District": "Causeway Bay",
        "Address": "Shop 124, 1/F Lee Garden Two, 28 Yun Ping Rd, CWB",
        "Phone": "+852 6683 3833",
        "Website / IG": "thecakery.com",
        "Booking": "Walk-in",
        "Hours": "10:00–20:00",
        "Price (HKD/pp)": "60–140",
        "Veg type": "Vegetarian (vegan + GF options)",
        "Jain": "Partial",
        "Signature / sample dishes (with prices where known)": "Sugar-free cakes; gluten-free cupcakes; vegan tarts; signature 'guilt-free' birthday cakes",
        "Google rating": "4.4 (250+)",
        "OpenRice rating": "—",
        "HappyCow rating": "4.4",
        "Michelin / Tatler / Time Out": "Tatler-listed",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Multi-branch HK (also Landmark/IFC). Very strong vegan dessert option.",
        "Sources (URLs)": "https://www.thecakery.com; Tatler Asia",
    },
    {
        "Name (EN)": "The Cakery (Landmark)",
        "Name (中文)": "—",
        "Sub-type": "Vegan / refined-sugar-free bakery",
        "District": "Central",
        "Address": "Shop 301, The Landmark, 15 Queen's Rd Central",
        "Phone": "+852 6683 3833",
        "Website / IG": "thecakery.com",
        "Booking": "Walk-in",
        "Hours": "10:00–20:00",
        "Price (HKD/pp)": "60–140",
        "Veg type": "Vegetarian (vegan + GF options)",
        "Jain": "Partial",
        "Signature / sample dishes (with prices where known)": "Sugar-free cakes; vegan cupcakes; chai-spice tart; matcha cake",
        "Google rating": "4.4 (250+)",
        "OpenRice rating": "—",
        "HappyCow rating": "4.4",
        "Michelin / Tatler / Time Out": "Tatler-listed",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Sister branch to Lee Garden — same menu, different traffic profile.",
        "Sources (URLs)": "https://www.thecakery.com",
    },
    {
        "Name (EN)": "Teakha",
        "Name (中文)": "茶。家",
        "Sub-type": "Tea house / cafe (veg-friendly)",
        "District": "Sheung Wan",
        "Address": "B, 18 Tai Ping Shan St, Sheung Wan",
        "Phone": "+852 2858 9185",
        "Website / IG": "teakha.com",
        "Booking": "Walk-in",
        "Hours": "11:00–19:00 (closed Tue)",
        "Price (HKD/pp)": "80–180",
        "Veg type": "Vegetarian-friendly (most of menu is veg)",
        "Jain": "Partial",
        "Signature / sample dishes (with prices where known)": "Single-origin teas; homemade scones; savoury veg toasts; cake of the day",
        "Google rating": "4.6 (500+)",
        "OpenRice rating": "—",
        "HappyCow rating": "4.4",
        "Michelin / Tatler / Time Out": "Tatler / Honeycombers feature",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Pioneered HK's craft-tea movement (Nana Chan). Mostly veg menu (some dairy).",
        "Sources (URLs)": "https://www.teakha.com; Tatler Asia",
    },
    {
        "Name (EN)": "Plantation Tea Bar",
        "Name (中文)": "—",
        "Sub-type": "Tea bar / cafe (veg-friendly)",
        "District": "Sai Ying Pun",
        "Address": "15 High St, Sai Ying Pun",
        "Phone": "+852 3607 0680",
        "Website / IG": "plantation.hk",
        "Booking": "Walk-in",
        "Hours": "11:00–19:00",
        "Price (HKD/pp)": "60–150",
        "Veg type": "Vegetarian-friendly",
        "Jain": "Partial",
        "Signature / sample dishes (with prices where known)": "Specialty teas; vegan-leaning pastries; verify vegan options before ordering",
        "Google rating": "4.4 (200+)",
        "OpenRice rating": "—",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "Tatler feature",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "By Nana Chan (Teakha founder). Focuses on craft tea + small bites.",
        "Sources (URLs)": "https://plantation.hk; Tatler Asia",
    },
]

# --- C. Mainstream w/ veg menu additions ---
C_NEW = [
    {
        "Name (EN)": "T'ang Court",
        "Name (中文)": "唐閣",
        "Sub-type": "Cantonese fine dining",
        "District": "Tsim Sha Tsui (Langham)",
        "Address": "1/F The Langham, 8 Peking Rd, TST",
        "Phone": "+852 2375 1133",
        "Website / IG": "langhamhotels.com",
        "Booking": "OpenTable, phone",
        "Hours": "Lunch + dinner",
        "Price (HKD/pp)": "800–1,500",
        "Veg type": "Vegetarian tasting menu available",
        "Jain": "On request",
        "Signature / sample dishes (with prices where known)": "Vegetarian tasting; mushroom-driven courses; veg dim sum on request",
        "Google rating": "4.6 (1,000+)",
        "OpenRice rating": "4.4",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "★★★ Michelin",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "One of HK's longest-running 3-star Cantonese restaurants. Pre-order veg tasting.",
        "Sources (URLs)": "https://www.langhamhotels.com; Michelin Guide",
    },
    {
        "Name (EN)": "Sun Tung Lok",
        "Name (中文)": "新同樂",
        "Sub-type": "Cantonese fine dining",
        "District": "Tsim Sha Tsui",
        "Address": "47 Kimberley Rd, TST (verify current location)",
        "Phone": "+852 2723 8222",
        "Website / IG": "suntunglok.com.hk",
        "Booking": "OpenTable, phone",
        "Hours": "Lunch + dinner",
        "Price (HKD/pp)": "400–1,200",
        "Veg type": "Veg-on-request; mushroom & tofu specialties",
        "Jain": "On request",
        "Signature / sample dishes (with prices where known)": "Claypot lettuce (veg version); veg dim sum; double-boiled vegetable soup",
        "Google rating": "4.4 (500+)",
        "OpenRice rating": "4.2",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "★★ Michelin",
        "Review-quality flag": "Trusted",
        "Operational status": "Open (verify current branch)",
        "Notes": "Has had multiple branches over the years. Confirm location before booking.",
        "Sources (URLs)": "https://www.suntunglok.com.hk; Michelin Guide",
    },
    {
        "Name (EN)": "Belon",
        "Name (中文)": "—",
        "Sub-type": "Modern French / European",
        "District": "Soho / Central",
        "Address": "41 Elgin St, Soho, Central (verify after recent chef changes)",
        "Phone": "+852 2152 2872",
        "Website / IG": "belonsoho.com",
        "Booking": "SevenRooms",
        "Hours": "Dinner only (verify)",
        "Price (HKD/pp)": "700–1,400",
        "Veg type": "Veg accommodation by request (48hr notice)",
        "Jain": "Limited",
        "Signature / sample dishes (with prices where known)": "Vegetarian course adaptations on request",
        "Google rating": "4.3 (300+)",
        "OpenRice rating": "—",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "★ Michelin (verify current status)",
        "Review-quality flag": "Mixed (chef changes)",
        "Operational status": "Verify",
        "Notes": "Has had ownership/chef shifts since 2020. Best to confirm operational status and veg accommodation by phone before booking.",
        "Sources (URLs)": "https://www.belonsoho.com; Michelin Guide archive",
    },
    {
        "Name (EN)": "ÉPURE",
        "Name (中文)": "—",
        "Sub-type": "Modern French fine dining",
        "District": "Tsim Sha Tsui (Harbour City)",
        "Address": "Shop 403, 4/F Ocean Terminal, Harbour City, TST",
        "Phone": "+852 3185 8338",
        "Website / IG": "epure.hk",
        "Booking": "OpenTable, phone",
        "Hours": "Lunch + dinner",
        "Price (HKD/pp)": "800–1,500",
        "Veg type": "Vegetarian tasting menu (on request)",
        "Jain": "On request",
        "Signature / sample dishes (with prices where known)": "7-course veg tasting; mushroom & root-vegetable plates; pre-dessert + dessert",
        "Google rating": "4.5 (300+)",
        "OpenRice rating": "4.3",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "★ Michelin",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Pierre Gagnaire group concept. Notify in advance for veg tasting.",
        "Sources (URLs)": "https://www.epure.hk; Michelin Guide",
    },
    {
        "Name (EN)": "Hansik Goo",
        "Name (中文)": "—",
        "Sub-type": "Modern Korean fine dining",
        "District": "Central",
        "Address": "1/F The Wellington, 198 Wellington St, Central (verify post-relocation)",
        "Phone": "+852 3611 1811",
        "Website / IG": "hansikgoo.hk",
        "Booking": "SevenRooms, phone",
        "Hours": "Lunch + dinner",
        "Price (HKD/pp)": "500–1,200",
        "Veg type": "Vegetarian options on request",
        "Jain": "Limited",
        "Signature / sample dishes (with prices where known)": "Vegetable banchan platter; mushroom dolsot; vegetable jeon",
        "Google rating": "4.5 (350+)",
        "OpenRice rating": "4.2",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "★ Michelin",
        "Review-quality flag": "Trusted",
        "Operational status": "Open (recently relocated — verify)",
        "Notes": "Chef Mingoo Kang (ex-Mingles Seoul). Sun-prepared/preserved veg highlights.",
        "Sources (URLs)": "https://www.hansikgoo.hk; Michelin Guide",
    },
    {
        "Name (EN)": "Chilli Fagara",
        "Name (中文)": "麻辣燙",
        "Sub-type": "Sichuan (with strong veg section)",
        "District": "Central",
        "Address": "7 Old Bailey St, Central",
        "Phone": "+852 2796 6866",
        "Website / IG": "chillifagara.com",
        "Booking": "Phone, walk-in",
        "Hours": "11:30–15:00; 17:00–23:00",
        "Price (HKD/pp)": "150–350",
        "Veg type": "Mixed (dedicated veg menu)",
        "Jain": "On request",
        "Signature / sample dishes (with prices where known)": "Hot-and-spicy veg pot ($218); spicy tofu; mapo tofu (veg); veg dan dan noodles",
        "Google rating": "4.5 (500+)",
        "OpenRice rating": "4.0",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "Michelin Recommended",
        "Review-quality flag": "Trusted (verify operating status — has had ownership shifts)",
        "Operational status": "Verify",
        "Notes": "Reserves one wok for veg-only cooking. Confirm operational status before going.",
        "Sources (URLs)": "https://chillifagara.com; Michelin Guide",
    },
    {
        "Name (EN)": "Paradise Dynasty (multiple)",
        "Name (中文)": "樂天皇朝",
        "Sub-type": "Shanghainese / dim sum",
        "District": "Multi-branch (IFC, Times Square, etc.)",
        "Address": "IFC Mall, Times Square CWB, Tai Po, Tuen Mun (verify branch list)",
        "Phone": "Branch-specific",
        "Website / IG": "paradisegp.com",
        "Booking": "OpenTable, walk-in",
        "Hours": "11:00–22:00",
        "Price (HKD/pp)": "100–250",
        "Veg type": "Mixed (multiple veg xiao long bao + veg dishes)",
        "Jain": "Limited",
        "Signature / sample dishes (with prices where known)": "8-flavour XLB (luffa is veg ~$78/8pcs); veg dumplings; veg dan dan noodles; mapo tofu",
        "Google rating": "4.2 (700+ aggregated)",
        "OpenRice rating": "4.0",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Singapore-origin chain. Reliable mid-range option with several genuinely veg XLB.",
        "Sources (URLs)": "https://www.paradisegp.com",
    },
]

# --- D. Indian / ME / Med additions ---
D_NEW = [
    {
        "Name (EN)": "Chutney Tandoor House",
        "Name (中文)": "—",
        "Sub-type": "Modern North Indian (mixed)",
        "District": "Central",
        "Address": "4/F Carfield Commercial Bldg, 75-77 Wyndham St, Central",
        "Phone": "—",
        "Website / IG": "chutneytandoorhouse.shop",
        "Booking": "SevenRooms",
        "Hours": "12:00–15:00; 18:30–23:00",
        "Price (HKD/pp)": "250–400",
        "Veg type": "Mixed (dedicated veg + Jain options)",
        "Jain": "Yes",
        "Signature / sample dishes (with prices where known)": "Aloo tikki chaat ($120); tandoori cauliflower ($180); kamala kofta curry ($230); palak paneer",
        "Google rating": "4.6 (400+)",
        "OpenRice rating": "4.3",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "Time Out HK feature",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Dedicates one tandoor to vegetarian cooking. Impossible Foods collaboration on some items.",
        "Sources (URLs)": "https://chutneytandoorhouse.shop; Time Out HK",
    },
    {
        "Name (EN)": "Khyber Pass Mess Club",
        "Name (中文)": "—",
        "Sub-type": "Indian (Chungking Mansions classic)",
        "District": "Tsim Sha Tsui",
        "Address": "Block E, 7/F Chungking Mansions, 36-44 Nathan Rd, TST",
        "Phone": "+852 2368 2387",
        "Website / IG": "—",
        "Booking": "Phone, walk-in",
        "Hours": "12:00–24:00",
        "Price (HKD/pp)": "150–300",
        "Veg type": "Mixed (significant veg section)",
        "Jain": "On request",
        "Signature / sample dishes (with prices where known)": "Saag paneer; paneer paratha; veg samosa; veg biryani",
        "Google rating": "4.3 (400+)",
        "OpenRice rating": "4.0",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "Trusted",
        "Operational status": "Open",
        "Notes": "Established 1991 — a Chungking Mansions institution. Classic North Indian with strong veg dishes.",
        "Sources (URLs)": "OpenRice; HappyCow",
    },
    {
        "Name (EN)": "Curry Leaf Indian Cuisine",
        "Name (中文)": "—",
        "Sub-type": "North + South Indian (mixed)",
        "District": "Jordan / Yau Ma Tei",
        "Address": "Temple St area, Jordan (verify exact address)",
        "Phone": "—",
        "Website / IG": "—",
        "Booking": "Phone, walk-in",
        "Hours": "Verify",
        "Price (HKD/pp)": "120–250",
        "Veg type": "Mixed",
        "Jain": "On request",
        "Signature / sample dishes (with prices where known)": "Veg curries; samosa chaat; pani puri",
        "Google rating": "4.1 (150+)",
        "OpenRice rating": "—",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "Verify",
        "Operational status": "Verify",
        "Notes": "Listed in HK veg-friendly resources but limited online presence. Phone before going.",
        "Sources (URLs)": "HappyCow",
    },
]

# --- Closures kept for reference ---
CLOSED_REF = [
    {
        "Name (EN)": "Soil to Soul",
        "Name (中文)": "—",
        "Sub-type": "Korean Buddhist temple food (was)",
        "District": "Tsim Sha Tsui",
        "Address": "K11 Musea, 17 Kowloon Rd, TST",
        "Phone": "—",
        "Website / IG": "soiltosoulhk.com",
        "Booking": "—",
        "Hours": "—",
        "Price (HKD/pp)": "—",
        "Veg type": "Vegan (was)",
        "Jain": "—",
        "Signature / sample dishes (with prices where known)": "(was) Sunim Temple Menu (~$780/person, 24 vegan dishes); fermented sides",
        "Google rating": "—",
        "OpenRice rating": "—",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "—",
        "Operational status": "PERMANENTLY CLOSED (~2024)",
        "Notes": "Famous Korean temple-food concept. Closed — included for reference so you don't search for it.",
        "Sources (URLs)": "Instagram archive; HappyCow",
    },
    {
        "Name (EN)": "Greenwoods Raw Cafe",
        "Name (中文)": "—",
        "Sub-type": "Raw vegan (was)",
        "District": "Tsim Sha Tsui",
        "Address": "13/F, 2 Carnarvon Rd, TST",
        "Phone": "—",
        "Website / IG": "—",
        "Booking": "—",
        "Hours": "—",
        "Price (HKD/pp)": "—",
        "Veg type": "Raw vegan (was)",
        "Jain": "—",
        "Signature / sample dishes (with prices where known)": "(was) Green smoothies; organic salads; vegan 'salmon' tartare",
        "Google rating": "—",
        "OpenRice rating": "—",
        "HappyCow rating": "—",
        "Michelin / Tatler / Time Out": "—",
        "Review-quality flag": "—",
        "Operational status": "CLOSED",
        "Notes": "HK's first raw-only cafe. Closed.",
        "Sources (URLs)": "HappyCow CLOSED listing",
    },
]

# ============================================================
# v3 ADDITIONS — discovered via April 2026 audit + research agents
# ============================================================

A_NEW2 = [
    {"Name (EN)": "Fat Chee Vegetarian", "Name (中文)": "—", "Sub-type": "Cantonese veg", "District": "Yuen Long", "Address": "G/F, 22 Fuk Hong St, Yuen Long", "Phone": "—", "Website / IG": "—", "Booking": "Walk-in", "Hours": "10:30–16:45, 17:30–21:45 (closed Tue)", "Price (HKD/pp)": "51–100", "Veg type": "Vegetarian", "Jain": "No", "Signature / sample dishes (with prices where known)": "Stir-fried veg, mock-meat with vegetables, daily soup, tofu hot pot", "Google rating": "—", "OpenRice rating": "3.9", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Long-running Yuen Long neighbourhood spot.", "Sources (URLs)": "OpenRice"},
    {"Name (EN)": "Po Lin Vegetarian (Yuen Long)", "Name (中文)": "普蓮素食", "Sub-type": "Cantonese veg", "District": "Yuen Long", "Address": "Shop 13A & 15, G/F, Block 2, Yuccie Square, 38 Yuen Long On Ning Rd", "Phone": "—", "Website / IG": "—", "Booking": "OpenRice, walk-in", "Hours": "11:00–22:00", "Price (HKD/pp)": "51–80", "Veg type": "Vegetarian", "Jain": "No", "Signature / sample dishes (with prices where known)": "Mock-meat dishes, tofu steak, vegetable stir-fries", "Google rating": "—", "OpenRice rating": "4.0", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Affordable Yuen Long mock-meat spot.", "Sources (URLs)": "OpenRice"},
    {"Name (EN)": "One Vegan Shop", "Name (中文)": "一素店", "Sub-type": "Vegan home-cooking takeaway", "District": "Yuen Long", "Address": "Shop 12, G/F, 31-35 Yuen Sun St, Yee Fung Bldg, Yuen Long", "Phone": "—", "Website / IG": "facebook.com/oneveganshop", "Booking": "Walk-in", "Hours": "Daily (verify hours)", "Price (HKD/pp)": "60–100", "Veg type": "Vegan", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Homemade desserts, vegan beverages, takeaway meals", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Trusted (low review volume)", "Operational status": "Open", "Notes": "Small homestyle vegan shop.", "Sources (URLs)": "Facebook"},
    {"Name (EN)": "Lotus Vegetarian (Tai Po)", "Name (中文)": "蓮花素食", "Sub-type": "Cantonese veg", "District": "Tai Po", "Address": "G/F, Tak Hong Bldg, 80 Kwong Fuk Rd, Tai Po", "Phone": "—", "Website / IG": "—", "Booking": "Walk-in", "Hours": "11:00–22:00", "Price (HKD/pp)": "60–120", "Veg type": "Vegetarian", "Jain": "No", "Signature / sample dishes (with prices where known)": "Traditional Cantonese veg dishes; mock-meat curry; clay-pot rice", "Google rating": "—", "OpenRice rating": "3.9", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Local Tai Po Cantonese veg.", "Sources (URLs)": "OpenRice"},
    {"Name (EN)": "Tin Tin Vegetarian", "Name (中文)": "天天素食", "Sub-type": "Cantonese veg + cart noodles", "District": "Sha Tin", "Address": "Shop 31A, 3/F, Sha Tin Plaza, 21-27 Sha Tin Centre St", "Phone": "—", "Website / IG": "—", "Booking": "Walk-in, Foodpanda delivery", "Hours": "Lunch + dinner", "Price (HKD/pp)": "70–130", "Veg type": "Vegetarian", "Jain": "No", "Signature / sample dishes (with prices where known)": "Cart noodles (build your own), veg dim sum, soups, stir-fries", "Google rating": "—", "OpenRice rating": "3.9", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Sha Tin Plaza branch with cart-noodle setup.", "Sources (URLs)": "Foodpanda; OpenRice"},
    {"Name (EN)": "Sam Bo Vegetarian", "Name (中文)": "三寶齋", "Sub-type": "Cantonese veg / dim sum", "District": "Sha Tin", "Address": "Shop 33, 3/F, Sha Tin Plaza, 21-27 Sha Tin Centre St", "Phone": "—", "Website / IG": "—", "Booking": "OpenRice, walk-in", "Hours": "11:00–22:00", "Price (HKD/pp)": "70–130", "Veg type": "Vegetarian", "Jain": "No", "Signature / sample dishes (with prices where known)": "Veg dim sum, mock-meat dishes, soup noodles", "Google rating": "—", "OpenRice rating": "3.9", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Long-running Sha Tin Plaza veg restaurant.", "Sources (URLs)": "OpenRice"},
    {"Name (EN)": "Mui Fat Buddhist Monastery Canteen", "Name (中文)": "妙法寺素食堂", "Sub-type": "Buddhist monastery 齋", "District": "Tuen Mun (Lam Tei)", "Address": "G/F, 18 Castle Peak Rd – Lam Tei, Tuen Mun", "Phone": "+852 2461 8567", "Website / IG": "—", "Booking": "Walk-in", "Hours": "09:00–19:00", "Price (HKD/pp)": "51–100", "Veg type": "Buddhist 齋 (no allium)", "Jain": "Yes (de facto)", "Signature / sample dishes (with prices where known)": "Set lunches, mock-meat curries, monastery-style soups", "Google rating": "—", "OpenRice rating": "3.9", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Inside the working monastery — dress respectfully.", "Sources (URLs)": "OpenRice"},
    {"Name (EN)": "Western Monastery Vegetarian Hall", "Name (中文)": "西方寺素食堂", "Sub-type": "Buddhist monastery 齋", "District": "Tsuen Wan", "Address": "Western Monastery, Lo Wai Rd, Tsuen Wan", "Phone": "—", "Website / IG": "—", "Booking": "Walk-in (visit during temple hours)", "Hours": "Daytime (verify)", "Price (HKD/pp)": "40–80", "Veg type": "Buddhist 齋 (no allium)", "Jain": "Yes (de facto)", "Signature / sample dishes (with prices where known)": "Simple monastery veg meals, congee, mock-meat dishes", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Discover Hong Kong listing", "Review-quality flag": "Trusted (low online presence)", "Operational status": "Open", "Notes": "Inside Western Monastery. Tied to temple visit.", "Sources (URLs)": "Discover Hong Kong"},
    {"Name (EN)": "Home Veggie", "Name (中文)": "素回家", "Sub-type": "Cantonese veg takeaway", "District": "Sha Tin (Yu Chui)", "Address": "Shop 2-3, G/F, Yu Chui Market, 2 Ngau Pei Sha St", "Phone": "—", "Website / IG": "—", "Booking": "WhatsApp pre-order", "Hours": "Daily (verify)", "Price (HKD/pp)": "50–80", "Veg type": "Vegetarian", "Jain": "No", "Signature / sample dishes (with prices where known)": "Pre-ordered rice meals, noodles, vegan desserts", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "—", "Review-quality flag": "Verify (low online presence)", "Operational status": "Verify", "Notes": "Mostly pre-order takeaway via WhatsApp.", "Sources (URLs)": "Search results"},
]

B_NEW2 = [
    {"Name (EN)": "Peridot", "Name (中文)": "—", "Sub-type": "Plant-based fine dining + cocktail bar", "District": "Central (The Henderson)", "Address": "Summit 38, 38/F, The Henderson, 2 Murray Rd, Central", "Phone": "—", "Website / IG": "peridothk.com", "Booking": "SevenRooms, phone", "Hours": "12:00–02:00 (Mon-Sat)", "Price (HKD/pp)": "500–800", "Veg type": "100% plant-based", "Jain": "On request", "Signature / sample dishes (with prices where known)": "Koji carrot pumpkin ginger soup; asado mushroom w/ chimichurri + chorizo tempeh; fleshy fruits cold cuts", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Time Out HK + SCMP feature; Tatler 2026 best new", "Review-quality flag": "Trusted", "Operational status": "Open (opened 2026)", "Notes": "Chef Lisandro Illa (Noma alum). Sky-high vegan fine dining at The Henderson. Studio Paolo Ferrari interior.", "Sources (URLs)": "https://peridothk.com; Time Out HK; SCMP"},
    {"Name (EN)": "Happy Samurai", "Name (中文)": "—", "Sub-type": "Vegan ramen", "District": "Central", "Address": "Shop A, G/F, 51 Wellington St, Central", "Phone": "—", "Website / IG": "happysamurai.hk", "Booking": "Walk-in", "Hours": "11:30–23:00 Mon-Sat (closed Sun)", "Price (HKD/pp)": "128–138", "Veg type": "Vegan", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Shoyu ($128); Tonkotsu ($138); Spicy Tantan ($138); all plant-based broths", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Time Out HK + SCMP feature", "Review-quality flag": "Trusted", "Operational status": "Open (April 2026)", "Notes": "Plant-based ramen — broth from 80-year-old Hokkaido soup-base factory.", "Sources (URLs)": "https://happysamurai.hk; Time Out HK"},
    {"Name (EN)": "Le FroMage by Ma", "Name (中文)": "法蒙芝士", "Sub-type": "Vegan cheese shop / counter", "District": "Tsim Sha Tsui (K11 Musea)", "Address": "Unit 21, B201, K11 MUSEA, Victoria Dockside, 18 Salisbury Rd, TST", "Phone": "—", "Website / IG": "lefromage.com.hk", "Booking": "Walk-in", "Hours": "10:00–22:00 daily", "Price (HKD/pp)": "120–180", "Veg type": "Vegan", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Shamembert (truffle); 'chiaviar' caviar-style; 'sans-salmon'; keto breads; gluten-free crackers", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Tatler + Green Queen", "Review-quality flag": "Trusted", "Operational status": "Open (March 2024)", "Notes": "HK's first 100% vegan cheese shop. Raw vegan chef Tina Barrat.", "Sources (URLs)": "Green Queen HK; Tatler Asia"},
    {"Name (EN)": "Dotom", "Name (中文)": "—", "Sub-type": "Korean plant-based kimbap", "District": "Central", "Address": "Shop 1, G/F & M/F, Welley Bldg, 97 Wellington St, Central", "Phone": "—", "Website / IG": "—", "Booking": "Walk-in", "Hours": "Daily (verify)", "Price (HKD/pp)": "80–120", "Veg type": "Vegetarian (vegan options)", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Classic kimbap; plant-protein noodle rolls; lettuce keto rolls; soba rolls; vegan inari", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Tatler 2026 best new", "Review-quality flag": "Trusted", "Operational status": "Open (2026)", "Notes": "Korean kimbap specialist — fermentation-focused, lots of veg options.", "Sources (URLs)": "Tatler Asia"},
    {"Name (EN)": "Maya Bakery & Bar", "Name (中文)": "—", "Sub-type": "Vegan artisan bakery", "District": "Quarry Bay (Taikoo Place) + Central (IFC)", "Address": "Taikoo: PCCW Tower, Taikoo Place | Central: IFC Mall, Finance St", "Phone": "—", "Website / IG": "mayabakerybar.com", "Booking": "Walk-in", "Hours": "Verify per branch", "Price (HKD/pp)": "40–120", "Veg type": "Vegan", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Vegan egg tarts; pineapple buns; pistachio croissants; blueberry muffins; PB-chocolate cookies", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Green Queen + The Beat", "Review-quality flag": "Trusted", "Operational status": "Open (2024-25)", "Notes": "Sister brand to The Cakery (Shirley Kwok). Almost fully plant-based.", "Sources (URLs)": "Green Queen HK; The Beat"},
    {"Name (EN)": "Bobsang", "Name (中文)": "—", "Sub-type": "Korean hot pot (large vegan section)", "District": "Tsing Yi (Maritime Square)", "Address": "Shop 119B, 1/F, Maritime Square, 33 Tsing King Rd, Tsing Yi", "Phone": "—", "Website / IG": "—", "Booking": "OpenRice", "Hours": "Verify", "Price (HKD/pp)": "100–150", "Veg type": "Vegetarian (Bobsang Green section is vegan)", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "20+ vegan DIY toppings (tteokbokki bowl); vegan banchan; soft serve (red dragonfruit, peach)", "Google rating": "—", "OpenRice rating": "3.8", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Honeycombers + SCMP feature", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Korean hot pot with strong dedicated vegan/veg section.", "Sources (URLs)": "Honeycombers; SCMP"},
]

C_NEW2 = [
    {"Name (EN)": "Terrace Boulud", "Name (中文)": "—", "Sub-type": "Modern French (with 'Le Potager' veg pathway)", "District": "Central", "Address": "25/F Landmark Prince's, 10 Chater Rd, Central", "Phone": "+852 2132 0066", "Website / IG": "mandarinoriental.com", "Booking": "OpenTable, phone", "Hours": "Lunch + dinner", "Price (HKD/pp)": "400–650", "Veg type": "Mixed; dedicated 'Le Potager' (vegetable garden) menu", "Jain": "On request", "Signature / sample dishes (with prices where known)": "Vegetable-centric seasonal dishes; 4 pathways (Tradition, Saison, Le Potager, Le Voyage)", "Google rating": "—", "OpenRice rating": "—", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Tatler 2026 + Mandarin Oriental", "Review-quality flag": "Trusted", "Operational status": "Open (March 2026)", "Notes": "Daniel Boulud's Asia debut. Rooftop with harbour views.", "Sources (URLs)": "Mandarin Oriental; Tatler Asia"},
    {"Name (EN)": "Salisterra", "Name (中文)": "—", "Sub-type": "Mediterranean", "District": "Admiralty (Upper House)", "Address": "49/F The Upper House, Pacific Place, 88 Queensway, Admiralty", "Phone": "+852 2918 1838", "Website / IG": "upperhouse.com", "Booking": "OpenTable", "Hours": "Lunch + dinner", "Price (HKD/pp)": "350–500", "Veg type": "Mixed; dedicated vegan/veg section", "Jain": "On request", "Signature / sample dishes (with prices where known)": "Roasted carrots w/ coconut yoghurt + cashew nuts; vegan pasta; seasonal veg plates", "Google rating": "4.5 (300+)", "OpenRice rating": "4.3", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "SCMP, Vogue HK, Time Out", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Chef Jun Tanaka. André Fu interior. Skyline views.", "Sources (URLs)": "Upper House; Time Out HK"},
    {"Name (EN)": "Bistrot du Vin", "Name (中文)": "—", "Sub-type": "French wine bar / bistrot", "District": "Sai Ying Pun", "Address": "165-166 Connaught Rd West, The Fine Wine Experience, Sai Ying Pun", "Phone": "—", "Website / IG": "—", "Booking": "Phone", "Hours": "Lunch + dinner (verify)", "Price (HKD/pp)": "180–280", "Veg type": "Mixed; à la carte veg options", "Jain": "Limited", "Signature / sample dishes (with prices where known)": "Mushroom linguini; salads; veg tartines; cheese boards", "Google rating": "—", "OpenRice rating": "4.0", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Tatler", "Review-quality flag": "Trusted", "Operational status": "Open (relocated March 2026)", "Notes": "Chef Guillaume Losguardi. Relocated from Kennedy Town in 2026.", "Sources (URLs)": "Time Out HK; Tatler Asia"},
]

D_NEW2 = [
    {"Name (EN)": "Omnia", "Name (中文)": "—", "Sub-type": "Lebanese / Levantine", "District": "Central (Soho)", "Address": "33 Staunton St, Soho, Central", "Phone": "—", "Website / IG": "omniarestaurant.com", "Booking": "Phone, OpenRice", "Hours": "Lunch + dinner (verify)", "Price (HKD/pp)": "120–200", "Veg type": "Mixed; full vegan/veg mezze section", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Black-chickpea hummus w/ portobello; tabbouleh; falafel; baba ghanoush; kibbeh variants", "Google rating": "—", "OpenRice rating": "4.1", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Bloomberg HK review", "Review-quality flag": "Trusted", "Operational status": "Open (May 2025)", "Notes": "Chef-owner Nadim Hamze (ex-Sumac). Two-level Levantine arches.", "Sources (URLs)": "Bloomberg; OpenRice"},
    {"Name (EN)": "Bushra", "Name (中文)": "—", "Sub-type": "Middle Eastern (Lebanese / Moroccan / Egyptian)", "District": "Tsim Sha Tsui East", "Address": "G6 & UG16, Tsim Sha Tsui Centre, 66 Mody Rd, TST East", "Phone": "—", "Website / IG": "bushra.com.hk", "Booking": "Phone", "Hours": "12:00–02:00 daily", "Price (HKD/pp)": "150–250", "Veg type": "Mixed; dedicated vegan/veg section", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Moutabal; falafel; ras el keba; roasted vegetables; gluten-free options", "Google rating": "—", "OpenRice rating": "4.0", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Time Out HK; HiveLife", "Review-quality flag": "Trusted", "Operational status": "Open", "Notes": "Live entertainment some nights.", "Sources (URLs)": "https://www.bushra.com.hk; Time Out HK"},
    {"Name (EN)": "Sumac", "Name (中文)": "—", "Sub-type": "Lebanese / Mediterranean", "District": "Central", "Address": "G/F, 8 Glenealy, Central", "Phone": "+852 9594 1173", "Website / IG": "sumacrestaurants.com", "Booking": "Phone", "Hours": "Lunch + dinner", "Price (HKD/pp)": "140–220", "Veg type": "Mixed; strong vegan/veg mezze section", "Jain": "Partial", "Signature / sample dishes (with prices where known)": "Mouhamara; falafel; batata harra; hummus variants; shawarma sides", "Google rating": "—", "OpenRice rating": "4.1", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "Tripadvisor; Wanderlog", "Review-quality flag": "Trusted", "Operational status": "Open (2024)", "Notes": "Chef Hussein Al-Ankoune. Authentic Lebanese ingredients.", "Sources (URLs)": "https://sumacrestaurants.com"},
    {"Name (EN)": "Mara", "Name (中文)": "—", "Sub-type": "Modern Mediterranean", "District": "Sheung Wan", "Address": "Shop 01, G/F Ovest, 77 Wing Lok St, Sheung Wan", "Phone": "—", "Website / IG": "marahk.com", "Booking": "Phone, OpenRice", "Hours": "Verify", "Price (HKD/pp)": "200–350", "Veg type": "Mixed; veg-friendly Mediterranean", "Jain": "Limited", "Signature / sample dishes (with prices where known)": "Roasted Mediterranean vegetables; olive oil-forward dishes; seasonal veg plates", "Google rating": "—", "OpenRice rating": "4.0", "HappyCow rating": "—", "Michelin / Tatler / Time Out": "DiningCity", "Review-quality flag": "Trusted", "Operational status": "Open (2025)", "Notes": "Contemporary Mediterranean focus on seasonality.", "Sources (URLs)": "https://marahk.com"},
]

CAT_A = list(CAT_A_BASE) + A_NEW + A_NEW2
CAT_B = list(CAT_B_BASE) + B_NEW + B_NEW2
CAT_C = list(CAT_C_BASE) + C_NEW + C_NEW2
CAT_D = list(CAT_D_BASE) + D_NEW + D_NEW2 + CLOSED_REF

# ============================================================
# QUALITY SCORING
# ============================================================

def parse_rating(s):
    """Parse a rating string like '4.4 (200+)' or '4.4' into (rating, count) or None."""
    if not s or s in ("—", "", "varies"):
        return None
    text = str(s).strip()
    m = re.match(r"([\d.]+)", text)
    if not m:
        return None
    try:
        rating = float(m.group(1))
    except ValueError:
        return None
    if not (1.0 <= rating <= 5.0):
        return None
    count_match = re.search(r"\(([\d,]+)", text)
    count = int(count_match.group(1).replace(",", "")) if count_match else 25
    return rating, count


def quality_score(r):
    """Return (score 0-100, tier letter).

    Combines:
      * Bayesian-shrunk weighted average across Google / OpenRice / HappyCow
      * Critic bonus (Michelin / Tatler / Time Out)
      * Review-quality penalty (mixed / suspect / verify)
      * Status penalty (closed / uncertain / verify)
    """
    PRIOR_MEAN = 3.85
    PRIOR_N = 30
    samples = []
    for key in ("Google rating", "OpenRice rating", "HappyCow rating"):
        parsed = parse_rating(r.get(key, ""))
        if parsed:
            rating, count = parsed
            shrunk = (rating * count + PRIOR_MEAN * PRIOR_N) / (count + PRIOR_N)
            samples.append((shrunk, count + PRIOR_N))

    if samples:
        total_w = sum(w for _, w in samples)
        agg = sum(s * w for s, w in samples) / total_w
    else:
        agg = PRIOR_MEAN  # no ratings — fall back to prior, will be penalised through critic absence

    # 1.0–5.0 → 0–100, anchored so 2.5 = 0 and 5.0 = 100
    base = max(0.0, min(100.0, ((agg - 2.5) / 2.5) * 100))

    # Critic bonus
    critic = (r.get("Michelin / Tatler / Time Out", "") or "").lower()
    bonus = 0
    if "★★★" in critic or "3 michelin" in critic or "3-star" in critic:
        bonus = 15
    elif "★★" in critic or "2 michelin" in critic or "2-star" in critic or "2 star" in critic:
        bonus = 10
    elif "★" in critic or "1 michelin" in critic or "1-star" in critic or "michelin" in critic:
        bonus = 7
    elif "bib" in critic or "tatler" in critic or "time out" in critic or "scmp" in critic or "bloomberg" in critic:
        bonus = 4

    # Review-trust penalty
    flag = (r.get("Review-quality flag", "") or "").lower()
    penalty = 0
    if "suspect" in flag:
        penalty -= 20
    elif "mixed" in flag:
        penalty -= 10
    if "verify" in flag:
        penalty -= 5

    score = base + bonus + penalty

    # Status penalty
    status = (r.get("Operational status", "") or "").lower()
    if "closed" in status:
        return 0.0, "—"
    if "uncertain" in status:
        score *= 0.6
    elif "verify" in status:
        score *= 0.85

    score = max(0.0, min(100.0, round(score, 1)))
    if score >= 85: tier = "S"
    elif score >= 75: tier = "A"
    elif score >= 65: tier = "B"
    elif score >= 55: tier = "C"
    elif score > 0:   tier = "D"
    else:             tier = "—"
    return score, tier


# Apply scores to every record
ALL = []
for cat_label, rows in [
    ("A. Chinese / Buddhist 齋", CAT_A),
    ("B. Modern Plant-Based / Fine", CAT_B),
    ("C. Mainstream w/ Strong Veg", CAT_C),
    ("D. Indian / ME / Med", CAT_D),
]:
    for r in rows:
        score, tier = quality_score(r)
        r["__score__"] = score
        r["__tier__"] = tier
        r["__cat__"] = cat_label
        ALL.append(r)

# ============================================================
# WORKBOOK
# ============================================================

COLUMNS = [
    ("Quality (0-100)", 12),
    ("Tier", 6),
    ("Name (EN)", 28),
    ("Name (中文)", 18),
    ("Category", 18),
    ("Sub-type", 22),
    ("District", 16),
    ("Address", 42),
    ("Phone", 18),
    ("Website / IG", 30),
    ("Booking", 20),
    ("Hours", 30),
    ("Price (HKD/pp)", 14),
    ("Veg type", 14),
    ("Jain", 8),
    ("Signature / sample dishes (with prices where known)", 60),
    ("Google rating", 14),
    ("OpenRice rating", 14),
    ("HappyCow rating", 14),
    ("Michelin / Tatler / Time Out", 26),
    ("Review-quality flag", 18),
    ("Operational status", 18),
    ("Notes", 50),
    ("Sources (URLs)", 60),
]

HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TIER_FONT = Font(name="Arial", bold=True, size=11)

CAT_FILLS = {
    "A. Chinese / Buddhist 齋": PatternFill("solid", start_color="FFF2CC"),
    "B. Modern Plant-Based / Fine": PatternFill("solid", start_color="E2EFDA"),
    "C. Mainstream w/ Strong Veg": PatternFill("solid", start_color="FCE4D6"),
    "D. Indian / ME / Med": PatternFill("solid", start_color="DDEBF7"),
}

TIER_FILLS = {
    "S": PatternFill("solid", start_color="2A7A4D"),
    "A": PatternFill("solid", start_color="6FBF87"),
    "B": PatternFill("solid", start_color="C7E8B5"),
    "C": PatternFill("solid", start_color="FFE9A8"),
    "D": PatternFill("solid", start_color="F4C7B8"),
    "—": PatternFill("solid", start_color="DDDDDD"),
}
TIER_FONT_COLOURS = {"S": "FFFFFF", "A": "1A3D24", "B": "1A3D24", "C": "5A4500", "D": "5A2200", "—": "555555"}

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

wb = Workbook()
wb.remove(wb.active)


def write_header(ws, header_row):
    for c, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[header_row].height = 38


def write_row(ws, r_idx, r, cat_label, cat_fill):
    values = [
        r["__score__"], r["__tier__"], r.get("Name (EN)", ""), r.get("Name (中文)", ""),
        cat_label, r.get("Sub-type", ""), r.get("District", ""), r.get("Address", ""),
        r.get("Phone", ""), r.get("Website / IG", ""), r.get("Booking", ""), r.get("Hours", ""),
        r.get("Price (HKD/pp)", ""), r.get("Veg type", ""), r.get("Jain", ""),
        r.get("Signature / sample dishes (with prices where known)", ""),
        r.get("Google rating", ""), r.get("OpenRice rating", ""), r.get("HappyCow rating", ""),
        r.get("Michelin / Tatler / Time Out", ""), r.get("Review-quality flag", ""),
        r.get("Operational status", ""), r.get("Notes", ""), r.get("Sources (URLs)", ""),
    ]
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=r_idx, column=c, value=val)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
        if c == 2:  # Tier
            cell.font = TIER_FONT
            cell.fill = TIER_FILLS.get(r["__tier__"], TIER_FILLS["—"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", bold=True, size=11, color=TIER_FONT_COLOURS.get(r["__tier__"], "000"))
        elif c == 1:
            cell.font = Font(name="Arial", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = cat_fill
        else:
            cell.font = BODY_FONT
            cell.fill = cat_fill


def add_sheet(title, rows, cat_label, summary_blurb=""):
    ws = wb.create_sheet(title=title)
    if summary_blurb:
        ws["A1"] = summary_blurb
        ws["A1"].font = Font(name="Arial", italic=True, size=10, color="555555")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        header_row = 2
    else:
        header_row = 1
    write_header(ws, header_row)
    rows_sorted = sorted(rows, key=lambda r: (-r["__score__"], r.get("Name (EN)", "")))
    cat_fill = CAT_FILLS[cat_label]
    for i, r in enumerate(rows_sorted):
        write_row(ws, header_row + 1 + i, r, cat_label, cat_fill)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(rows_sorted)}"
    return ws


# ----- Sheets -----

# Top Picks (sorted by score across all categories)
top_picks = wb.create_sheet(title="★ Top Picks (by quality)")
top_picks["A1"] = "All restaurants ranked by composite Quality Score (0-100). Tier: S=85+, A=75-84, B=65-74, C=55-64, D<55. Closed/zero-score entries sink to bottom."
top_picks["A1"].font = Font(name="Arial", italic=True, size=10, color="555555")
top_picks.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
write_header(top_picks, 2)
all_sorted = sorted(ALL, key=lambda r: (-r["__score__"], r.get("Name (EN)", "")))
for i, r in enumerate(all_sorted):
    write_row(top_picks, 3 + i, r, r["__cat__"], CAT_FILLS[r["__cat__"]])
top_picks.freeze_panes = top_picks.cell(row=3, column=4)
top_picks.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(all_sorted)}"

# Master (alphabetical/category-grouped — also score-aware)
master = wb.create_sheet(title="MASTER (all categories)")
master["A1"] = "Master sheet — all four categories. Sortable by Quality Score (column A) or Tier (column B) using the AutoFilter dropdowns."
master["A1"].font = Font(name="Arial", italic=True, size=10, color="555555")
master.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
write_header(master, 2)
for i, r in enumerate(all_sorted):
    write_row(master, 3 + i, r, r["__cat__"], CAT_FILLS[r["__cat__"]])
master.freeze_panes = master.cell(row=3, column=4)
master.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + len(all_sorted)}"

# Per-category sheets
add_sheet("A. Chinese & Buddhist 齋", CAT_A, "A. Chinese / Buddhist 齋",
          "Dedicated Chinese vegetarian, Buddhist 齋, mock-meat, traditional dim sum. Sorted by Quality Score desc.")
add_sheet("B. Modern Plant-Based & Fine", CAT_B, "B. Modern Plant-Based / Fine",
          "Newer-wave dedicated vegan/vegetarian, modern fine-dining veg, plant-based cafes & fast-casual. Sorted by Quality Score desc.")
add_sheet("C. Mainstream w. Veg Menu", CAT_C, "C. Mainstream w/ Strong Veg",
          "Non-vegetarian restaurants with substantial dedicated veg menus / tasting menus. Sorted by Quality Score desc.")
add_sheet("D. Indian-ME-Med", CAT_D, "D. Indian / ME / Med",
          "Indian (incl. pure-veg), Middle Eastern, and Mediterranean restaurants with strong veg/Jain options. Sorted by Quality Score desc.")

# README
readme = wb.create_sheet(title="README & methodology")
notes = [
    ("Hong Kong Vegetarian Restaurants Database — v2", True),
    (f"Compiled April 2026. {len(ALL)} entries across four categories.", False),
    ("", False),
    ("How the Quality Score works", True),
    ("Each entry is scored 0-100 by combining:", False),
    ("  1. Weighted user ratings (Google, OpenRice, HappyCow) — Bayesian-shrunk so a 4.9★ with 8 reviews doesn't beat a 4.5★ with 800.", False),
    ("  2. Critic bonus: +15 for ★★★ Michelin, +10 for ★★, +7 for ★/Michelin Selected, +4 for Bib/Tatler/Time Out/SCMP/Bloomberg coverage.", False),
    ("  3. Review-trust penalty: -10 for 'mixed signals' (bimodal score patterns), -20 for 'suspect', -5 for 'verify'.", False),
    ("  4. Status penalty: closed = 0, uncertain × 0.6, verify × 0.85.", False),
    ("", False),
    ("Tier mapping", True),
    ("  S (85+) — top tier; ★★★ Michelin or universally praised standouts.", False),
    ("  A (75-84) — strong picks; serious destination places.", False),
    ("  B (65-74) — solid; reliable choice if it fits your need.", False),
    ("  C (55-64) — good but with caveats (mixed reviews, limited info, niche).", False),
    ("  D (<55) — significant caveats (suspect reviews, sparse data, uncertain status).", False),
    ("  — = closed.", False),
    ("", False),
    ("How to use the workbook", True),
    ("  • ★ Top Picks tab: all places ranked, S→D. Best entry point.", False),
    ("  • MASTER tab: same data, also score-sorted; use AutoFilter on Tier/Category/District/Veg type for slicing.", False),
    ("  • Per-category tabs: easier to scan one category at a time.", False),
    ("  • Colour coding: Yellow=Chinese, Green=Plant-based, Peach=Mainstream, Blue=Indian/ME/Med.", False),
    ("", False),
    ("Watching for paid reviews / bimodal scoring", True),
    ("OpenRice in particular shows bimodal score patterns at certain restaurants (clusters of 5★ alongside 1★ with hollow middle).", False),
    ("• Paradise Veggie 樂園素食 — flagged 'mixed signals' for this reason.", False),
    ("• Belon — also 'mixed' (chef changes + lower review volume after).", False),
    ("• Veggie Palace — 'mixed' due to very low review volume.", False),
    ("• HappyCow tends upward-biased — vegan users rate generously. Critic recognition is weighted independently.", False),
    ("", False),
    ("Operational status & closures", True),
    ("HK F&B has had heavy churn 2020-2026. Confirmed closures kept for reference (so you don't search for them in vain):", False),
    ("• Branto Pure Veg Indian (TST) — closed July 2025.", False),
    ("• Bedu (Sheung Wan, Middle Eastern) — closed June 2025.", False),
    ("• Soil to Soul (TST, Korean temple food) — closed ~2024.", False),
    ("• Greenwoods Raw Cafe — closed.", False),
    ("• Veda Café (Ovolo Central) — UNCERTAIN: website live but Instagram has indicated permanently closed. Call before going.", False),
    ("• Loving Hut HK — chain has shrunk; verify branches on lovinghut.com.", False),
    ("", False),
    ("Caveats", True),
    ("Some data points (exact phones, current hours, OpenRice review counts) shift frequently. The workbook captures best-available data at compile time but for high-stakes plans (anniversaries, group bookings) confirm by phone or via Instagram before going.", False),
    ("Ratings shown as 'X.X (N+ reviews)' are aggregated/rounded; counts shift weekly. Use as relative signal, not exact truth.", False),
    ("", False),
    (f"Total entries: {len(ALL)} ({len([r for r in ALL if r['__score__'] > 0])} active + {len([r for r in ALL if r['__score__'] == 0])} closed-for-reference).", True),
]
for i, (text, bold) in enumerate(notes, start=1):
    cell = readme.cell(row=i, column=1, value=text)
    cell.font = Font(name="Arial", bold=bold, size=12 if bold else 10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
readme.column_dimensions["A"].width = 130

# Reorder
wb._sheets = [
    wb["★ Top Picks (by quality)"],
    wb["MASTER (all categories)"],
    wb["A. Chinese & Buddhist 齋"],
    wb["B. Modern Plant-Based & Fine"],
    wb["C. Mainstream w. Veg Menu"],
    wb["D. Indian-ME-Med"],
    wb["README & methodology"],
]

wb.save(OUT)

# Quick self-report
tier_counts = {}
for r in ALL:
    tier_counts[r["__tier__"]] = tier_counts.get(r["__tier__"], 0) + 1
print(f"Saved: {OUT}")
print(f"Total entries: {len(ALL)}")
print(f"Tier counts: {sorted(tier_counts.items())}")
print("\nTop 10:")
for r in all_sorted[:10]:
    print(f"  {r['__score__']:5.1f}  {r['__tier__']}  {r.get('Name (EN)', '')[:40]:<40} | {r['__cat__']}")
print("\nBottom 5 (active):")
for r in [x for x in all_sorted if x['__score__'] > 0][-5:]:
    print(f"  {r['__score__']:5.1f}  {r['__tier__']}  {r.get('Name (EN)', '')[:40]:<40} | {r['__cat__']}")
